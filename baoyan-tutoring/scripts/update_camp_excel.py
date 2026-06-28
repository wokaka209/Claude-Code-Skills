#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
保研夏令营 Excel 主表批量更新模板脚本

用途：将核验结果按统一列映射批量写入 Excel 主表，生成版本化交付物。
使用前：修改 SRC/DST 路径与 updates 字典为本次核验结果。
运行：PYTHONUTF8=1 python update_camp_excel.py

列映射:
  A(1)=层次  B(2)=学校  C(3)=学院  D(4)=类别
  E(5)=截止时间  F(6)=状态  G(7)=备注
  H(8)=通知链接  I(9)=通知描述  J(10)=可报名标记  K(11)=核验日期

数据从行 4 开始（行 2 为口径说明行）。
"""
import openpyxl
import shutil
import os

# ============================================================
# 路径配置：每次使用时更新
# ============================================================
SRC = r"E:\Workbuddy_workplace\保研辅导\outputs\YYYY-MM-DD_camp_*_update\2027届自动化专业保研夏令营信息汇总.xlsx"
DST = r"E:\Workbuddy_workplace\保研辅导\outputs\YYYY-MM-DD_camp_status_update\2027届自动化专业保研夏令营信息汇总.xlsx"
SHEET_NAME = '2027自动化夏令营总表'
VERIFY_DATE = 'YYYY-MM-DD'  # 本次核验日期

# 复制上一版本作为基础
os.makedirs(os.path.dirname(DST), exist_ok=True)
shutil.copy2(SRC, DST)

wb = openpyxl.load_workbook(DST, data_only=False)
ws = wb[SHEET_NAME]

# ============================================================
# 更新数据字典: 行号 -> {列号: 新值}
# 列: 5=截止时间(E) 6=状态(F) 7=备注(G) 8=通知链接(H) 9=通知描述(I) 10=可报名标记(J) 11=核验日期(K)
# None 表示清空该列
# ============================================================
updates = {}

# ------ 示例：A类 截止时间确认/更新 + 状态变化 ------
# updates[行号] = {
#     5: '学院名~2026-07-01 17:00',
#     6: '已开通（学院名）',
#     7: '详细备注：通知发布日期、截止时间、活动时间、申请条件、命名差异、与自动化方向相关性。',
#     8: 'https://xxx.edu.cn/notice',
#     9: '一句话概括：通知发布与截止情况。',
#     10: '可报名：学院名7/1 17:00截止',
#     11: VERIFY_DATE,
# }

# ------ 示例：B类 已截止 ------
# updates[行号] = {
#     6: '已截止（学院名6/25截止）',
#     7: '学院名夏令营6/25截止(确认)，面向2027届。可关注预推免。',
#     9: '学院名6/25截止(确认)。',
#     10: '已截止',
#     11: VERIFY_DATE,
# }

# ------ 示例：C类 仍未见通知 ------
# updates[行号] = {
#     6: f'{VERIFY_DATE}核验仍未见2027届通知',
#     7: '学院官网最新为xxx，无夏令营通知。',
#     9: f'{学校名}：{VERIFY_DATE}核验仍未见2027届自动化方向夏令营通知。',
#     10: None,  # 清空可报名标记
#     11: VERIFY_DATE,
# }

# ------ 示例：D类 无变化仅更新核验日期 ------
# updates[行号] = {
#     11: VERIFY_DATE,
#     9: f'{学校名}：{VERIFY_DATE}核验无变化。',
# }

# ============================================================
# 执行更新
# ============================================================
updated_count = 0
for row, cols in updates.items():
    school = ws.cell(row=row, column=2).value
    for col, val in cols.items():
        if val is not None:
            ws.cell(row=row, column=col).value = val
        else:
            ws.cell(row=row, column=col).value = None
    updated_count += 1
    print(f'  Row {row} ({school}): 已更新 {len(cols)} 列')

# ============================================================
# 更新口径说明行 (Row 2)
# ============================================================
ws.cell(row=2, column=1).value = (
    f'口径：覆盖自动化、控制科学与工程、人工智能、电子信息、电气、智能制造等强相关方向；'
    f'全量核验至{VERIFY_DATE}，共核验{updated_count}所院校；'
    f'截止时间带"约"字为聚合平台倒计时换算，以官网为准。'
    f'【{VERIFY_DATE}重点变化】（在此填写本次核验的新增通知/截止变化/已截止摘要）。'
)

wb.save(DST)
print(f'\n更新完成: 共更新 {updated_count} 行')
print(f'输出文件: {DST}')
